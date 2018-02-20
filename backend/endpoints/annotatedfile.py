import falcon
from endpoints.baserequest import BaseRequest
from common import ns_lada, ns_data, ns_cube, ns_lcd, ns_dct, ns_xsd, ns_qb4cc, ns_skos
from logs import main_logger, app_logger

from rdflib.graph import Graph
from rdflib import Namespace, Literal, URIRef, BNode
from rdflib.namespace import RDF, FOAF, XSD, RDFS, OWL, SKOS

from openpyxl import *
from openpyxl.utils import *

import json
import requests
import urllib

import io
import os
import uuid
import mimetypes
import sys, traceback

from tablemapper import TableMapper

class AnnotatedFilesJSON(BaseRequest):

    def on_get(self, req, resp):
        uri = req.get_param('uri')
        sheetName = req.get_param('sheet')
        filePath = self.dm.value(URIRef(uri), ns_lada['filePath'], 'gpubs')
        main_logger.info(filePath)
        wb = load_workbook(filename = filePath, data_only=True)
        sheet = wb[sheetName]

        data = []
        isFirst = True
        for row in sheet.iter_rows():
            if (isFirst):
                isFirst = False
                continue
            if len(row) > 0:
                rowData = []
                for cell in row:
                    if(cell.value != None and (cell.value == 'file:' or cell.value == 'comment:' or cell.value == 'label:')):
                        break
                    else:
                        if (cell.value == None):
                            rowData.append('')
                        else:
                            rowData.append(cell.value)
                if(len(rowData) > 0):
                    data.append(rowData)


        resp.status = falcon.HTTP_200
        resp.content_type = 'application/json'
        resp.data = json.dumps(data)


class AnnotatedFiles(BaseRequest):
    _CHUNK_SIZE_BYTES = 4096


    def transformResources(self, graphURI, sourceProperty, targetType):
        #Expressions
        query = """
            SELECT ?ds ?prop ?obs ?label {
                    ?obs qb:dataSet ?ds .
                    ?obs %s ?prop .
                    ?prop rdfs:label ?label

            }""" % (sourceProperty)
        #graph = g.get_context(graphURI)
        qres = self.dm.query(query, [graphURI, 'glcd'])
        self.dm.export(graphURI)
        ggen = Graph()
        for row in qres:
            obs = row.obs
            dsURI = row.ds
            prop = row.prop
            #main_logger.info(prop);
            label = row.label
            try:
                propURI = URIRef(prop)
                ggen.add( (obs, ns_lada[targetType], propURI ))
                ggen.add( (propURI, RDFS.label, label) )
            except KeyError:
                main_logger.error("Could not add " + targetType + " URI for value " + prop)

        ggen.serialize('gen-cor.ttl', format='turtle')
        self.dm.add_graph(ggen, 'ggen')

    def transformLiterals(self, graphURI, sourceProperty, targetType):
        #Expressions
        query = """
            SELECT ?ds ?prop ?obs {
                    ?obs qb:dataSet ?ds .
                    ?obs %s ?prop

            }""" % (sourceProperty)
        qres = self.dm.query(query, graphURI)
        ggen = Graph()
        for row in qres:
            obs = row.obs
            dsURI = row.ds
            prop = row.prop
            try:
                propURI = URIRef(dsURI + '/' + targetType + '/' + urllib.quote_plus(prop))
                ggen.add( (obs, ns_lada['gen_' + targetType], propURI) )
                ggen.add( (propURI, RDFS.label, prop) )
            except KeyError:
                main_logger.error("Could not add " + targetType + " URI for value " + prop)

        #ggen.serialize("ggen.ttl", format="turtle")
        self.dm.add_graph(ggen, 'ggen')

    def transformPeriods(self, graphURI, targetType):
        #Expressions
        query = """
            SELECT ?ds ?obs ?startYear ?endYear {
                    ?obs qb:dataSet ?ds .
                    ?obs <http://h224.it.helsinki.fi:8080/varieng/data/startYear> ?startYear .
                    ?obs <http://h224.it.helsinki.fi:8080/varieng/data/endYear> ?endYear .

            }"""

        qres = self.dm.query(query, graphURI)
        ggen = Graph()
        for row in qres:
            obs = row.obs
            dsURI = row.ds
            startYear = row.startYear
            endYear = row.endYear
            try:
                propURI = URIRef(dsURI + '/' + targetType + '/' + str(startYear) + '-' + str(endYear))
                ggen.add( (obs, ns_lada[targetType], propURI) )
                ggen.add( (propURI, RDFS.label, Literal(str(startYear) + '-' + str(endYear))) )

            except KeyError:
                main_logger.error("Could not add " + targetType + " URI for value " + prop)

        #ggen.serialize("ggen-period.ttl", format="turtle")
        self.dm.add_graph(ggen, 'ggen')

    def save_file(self, file):
        path = os.path.join(self.dm.get_file_storage(), str(file.filename).replace(' ', '_'))
        with io.open(path, 'wb') as fileOut:
            while True:
                chunk = file.file.read(self._CHUNK_SIZE_BYTES)
                if not chunk:
                    break
                fileOut.write(chunk)
        return path

    def store_publication_metadata(self, data, filePath, rdfPath):
        gpubs = Graph()
        pub = URIRef(data['uri'])
        gpubs.add( (pub, RDF.type, ns_lcd['Publication']) )
        gpubs.add( (pub, ns_dct['title'], Literal(data['title'])) )
        gpubs.add( (pub, ns_dct['issued'], Literal(data['year'])) )
        gpubs.add( (pub, ns_dct['creator'], Literal(data['authors'])) )
        gpubs.add( (pub, ns_lada['filePath'], Literal(filePath)) )
        gpubs.add( (pub, ns_lada['rdfPath'], Literal(rdfPath)) )
        gpubs.add( (pub, ns_lada['file'], URIRef(data['fileURI'])) )

        self.dm.add_graph(gpubs, 'gpubs')


    def get_publication_metadata(self, graphURI):
        # query cube data for file URI
        queryStr = "select ?uri { GRAPH <" + graphURI + "> { ?ds a <http://purl.org/linked-data/cube#DataSet> . ?ds <http://data.hulib.helsinki.fi/id/ontology/varieng/v1/file> ?uri } }"
        print queryStr

        qres = self.dm.query_all(queryStr)
        for row in qres:
            fileURI = row['uri']

        if self.checkForConnection():
            resp = requests.get(
                'http://' + self.dm.lcdURL + ':' + str(self.dm.lcdPort) + '/varieng/queries/public-publication-from-file.rq?results&tqx=out:sparql-json&uri=<{0}>'.format(fileURI)
            )
            if(resp.ok):
                data = json.loads(resp.content)
                results = data['results']['bindings']
                if(len(results) > 0):
                    b = results[0]
                    return {
                        'uri': b['pub']['value'],
                        'title': b['title']['value'],
                        'year': b['year']['value'],
                        'authors': b['authors']['value'],
                        'fileURI': fileURI
                    }
            else:
                resp.raise_for_status()

        else:
            _uuid = str(uuid.uuid4())
            return {
                'uri': 'http://lada/demo/pub' + _uuid,
                'title': 'publication ' + _uuid,
                'year': 'Unknown',
                'authors': 'Unknown',
                'fileURI': fileURI
            }

    def generate_rdf(self, path):
        filename = os.path.basename(path)
        rdfDir = str(path).replace(' ', '_')
        outputFolder = os.path.join(path, rdfDir + "-rdf")
        print "Creating dir: " + outputFolder
        try:
            os.makedirs(outputFolder)
        except OSError:
            if not os.path.isdir(outputFolder):
                raise

        try:
            wb = load_workbook(filename = path)
            for sheet in wb.worksheets:
                print sheet.title
                mapper = TableMapper()
                mapper.initSheet(sheet)
                if not(mapper.isSkipped()):
                    try:
                        prefix =  rdfDir + '-' + sheet.title + '-o-'
                        g = mapper.generateGraph(prefix)
                        outFile = outputFolder + '/' + filename + '-' + sheet.title + '.ttl'
                        g.serialize(outFile, format="turtle")
                        print "Writing " + filename + ".ttl"
                        #print g.serialize(format='nt')

                    except:
                        exc_type, exc_value, exc_traceback = sys.exc_info()
                        traceback.print_exception(exc_type, exc_value, exc_traceback, limit=5, file=sys.stdout)

                        print "Could not create dataset description for file " + filename
        except:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            traceback.print_exception(exc_type, exc_value, exc_traceback, limit=5, file=sys.stdout)
            print "Error generating RDF from file " + path

        return outputFolder

    def store_rdf(self, inputFolder):
        graphURI = URIRef(':' + inputFolder)
        for infile in os.listdir(inputFolder):
            main_logger.info('File:' + str(infile))
            self.dm.add_file(inputFolder +'/'+ infile, 'turtle', graphURI)

        #graph.serialize("graph.ttl", format="turtle")
        return graphURI

    def on_post(self, req, resp):
        data = {
            'files': []
        }
        filenames = req.params
        for filename in filenames:
            try:
                main_logger.info(filename)
                file = req.get_param(filename)
                path = self.save_file(file)
                rdfPath = self.generate_rdf(path)
                graphURI = self.store_rdf(rdfPath)
                self.transformLiterals(graphURI, 'data:Expression', 'expression')
                self.transformLiterals(graphURI, 'data:Function', 'function')
                self.transformLiterals(graphURI, 'data:Genre', 'genre')
                self.transformPeriods(graphURI, 'timeperiod')
                self.transformResources(graphURI, 'data:Corpus', 'corpus')

                self.dm.export('ggen')
                pubMetadata = self.get_publication_metadata(graphURI)

                self.store_publication_metadata(pubMetadata, path, rdfPath)


                data['files'].append(
                    {
                        'filename': filename,
                        'path': path,
                        'uri': str(graphURI),
                        'status': 'OK'
                    }
                );
            except:
                exc_type, exc_value, exc_traceback = sys.exc_info()
                traceback.print_exception(exc_type, exc_value, exc_traceback, limit=5, file=sys.stdout)

                data['files'].append(
                    {
                        'filename': filename,
                        'path': path,
                        'uri': str(graphURI),
                        'status': 'ERROR',
                        'message': 'change me'
                    }
                );
        #gpubs.serialize("pubs.ttl", format="turtle")
        #g.serialize("g.ttl", format="turtle")

        resp.status = falcon.HTTP_200
        resp.content_type = 'application/json'
        resp.data = json.dumps(data)

    def on_delete(self, req, resp):
        main_logger.info('Delete: annotatedFiles')
        uri = req.get_param('uri')
        folder = req.get_param('file')
        # delete rdf folder

        # delete annotated file

        # remove rdf from the pubs graph
        self.dm.remove( (URIRef(uri), None, None), None )

        # remove cube data
        cubeGraphURI = URIRef(':' + folder)
        main_logger.info(cubeGraphURI)
        self.dm.clear(cubeGraphURI)

        #g.serialize("g.ttl", format="turtle")
        resp.status = falcon.HTTP_200
        resp.content_type = 'application/json'
