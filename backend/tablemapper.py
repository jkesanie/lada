from rdflib import URIRef, BNode, Literal
from rdflib.namespace import RDF, FOAF, XSD, RDFS
from rdflib import Graph
from openpyxl import *
from openpyxl.utils import *
import uuid
import sys, traceback
import os
import urllib
import pprint
import re
import regex
import json
from openpyxl.styles import NamedStyle, Font, Border, Side, Color
from openpyxl import styles

baseURL = 'http://h224.it.helsinki.fi:8080/varieng/data/'

corpusBaseURL = 'http://h224.it.helsinki.fi:8080/varieng/data/corpus/'
genreBaseURL = 'http://h224.it.helsinki.fi:8080/varieng/data/genre/'

vg_startYear = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/startYear")
vg_endYear = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/endYear")
vg_freq = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/data/frequency")
vg_expression = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/expression")
vg_function = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/function")

vg_grammar = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/grammar")
vg_genre = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/genre")

vg_corpus = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/corpus")
vg_timeperiod = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/timeperiod")
vg_text = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/text")
vg_per = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/data/per")


vg_file = URIRef("http://data.hulib.helsinki.fi/id/ontology/varieng/v1/file")

qb_obs = URIRef("http://purl.org/linked-data/cube#Observation")
qb_dataSet = URIRef("http://purl.org/linked-data/cube#dataSet")
qb_DataSet = URIRef("http://purl.org/linked-data/cube#DataSet")

lada_row = URIRef('http://lada/row')
lada_col = URIRef('http://lada/col')
lada_sheet = URIRef('http://lada/sheet')


pp = pprint.PrettyPrinter(indent=4)

#timePeriodRegex = re.compile(r'(\d\d\d\d).{0,2}(\d\d\d\d)',re.UNICODE)
timePeriodRegex = re.compile(ur'(\d\d\d\d)-(\d\d\d\d)',re.UNICODE)
timePeriodRegex2 = re.compile(ur'(\d\d\d\d)s',re.UNICODE) # 1960x
combinedNumberRegex = re.compile(r'(\d+)\s*\((\d+\.\d+|\d+)\)') # x or x (y.z)

#corpusMapping = None
#with open('/Users/jkesa/Documents/LCD/LaDa/data/corpus-mapping.json') as data_file:
    #corpusMapping = json.load(data_file)

class TableMapper:

    cells = None
    values = None
    common = None
    dataset = None
    maxRow = -1
    maxCol = -1

    def __init__(self):
        self.values = []
        self.common = []
        self.dataset = []
        self.maxCol = -1
        self.maxRow = -1
        self.cells = None

    def getAnnotationType(self, cell):
        # annotation type is determined by the color of the cell
        valueColor = 'FFDDDDDD'
        corpusColor = 'FF9BC995'
        periodColor = 'FF98B9AB'
        expressionColor = 'FF5171A5'
        functionColor = 'FFC2AFF0'
        genreColor = 'FFEEF36A'


        color = cell.fill.fgColor.rgb
        if color == corpusColor:
            return 'Corpus'
        if color == periodColor:
            return 'Period'
        if color == expressionColor:
            return 'Expression'
        if color == functionColor:
            return 'Function'
        if color == genreColor:
            return 'Genre'
        if color == valueColor:
            return 'Value'
        return cell.style

    def getValueAndTypeWithMergeLookup(self,sheet, cell):
        if cell == None or sheet == None:
            return None
        for irange in sheet.merged_cell_ranges:
            min_col, min_row, max_col, max_row =range_boundaries(irange)
            if cell.row in range(min_row,max_row+1) and column_index_from_string(cell.column) in range(min_col,max_col+1):
                cell = sheet.cell(None,min_row,min_col)
                value =  cell.value
                atype = self.getAnnotationType(cell)
                return (value, atype)
        value =  cell.value
        atype = self.getAnnotationType(cell)
        return (value, atype)

    def traverseCells(self, sheet):

        maxRow =  sheet.max_row
        maxCol = sheet.max_column
        self.maxCol = maxCol
        self.maxRow = maxRow
        #print str(maxCol) + "," + str(maxRow)
        self.cells = [[1 for x in range(maxRow + 1)] for y in range(maxCol + 1)]
        for rowIndex in range(1, maxRow + 1):
            for colIndex in range(1, maxCol + 1):
                #print "(" + str(colIndex) + ", " + str(rowIndex) + ")"
                cell = sheet.cell(row = rowIndex, column = colIndex)
                    #print unicode(cell.value).encode('utf-8', 'ignore')
                isURI = False
                if cell.font.underline != None:
                    isURI = True


                annotationType = self.getAnnotationType(cell)
                annotationValue = cell.style

                value = cell.value

                cellName = cell.column + str(cell.row)
                if value == None and cellName in sheet.merged_cells:
                    values = self.getValueAndTypeWithMergeLookup(sheet, cell)
                    value = values[0]
                    annotationType = values[1]


                isCommon = cell.font.bold and annotationType != 'Normal' and annotationType != 'Value'

                isValue = False
                if "Value" == annotationType:
                    #print "value = " + unicode(cell.value).encode('utf-8', 'ignore')
                    isValue = True


                obj = []
                if 'Corpus' == annotationType:
                    # TODO: mapping to URIRef
                    #corpusURI = corpusMapping[annotationValue]
                    #obj.append( { 'value' : annotationValue, 'avalue': annotationValue, 'type' : annotationType, 'isValue': isValue, 'isURI': False, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )
                    # TODO: get rid of this !

                    av =  annotationValue.lower()
                    if av == 'bnc written':
                        av = 'british+national+corpus+-+written'
                    elif av == 'bnc spoken':
                        av = 'british+national+corpus+-+spoken'
                    elif av == 'med':
                        av = 'middle+english+dictionary'

                    obj.append( { 'value' : 'http://h224.it.helsinki.fi:8080/varieng/data/corpus/' + av, 'avalue': av, 'type' : annotationType, 'isValue': isValue, 'isURI': True, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )

                elif 'Period' == annotationType:
                    #main_logger.info('PERIOD')
                    # range
                    #valueMatches = timePeriodRegex.match(unicode(value))
                    valueMatches = regex.search(ur'(\d\d\d\d).{0,1}(\d\d\d\d)', unicode(value))
                    valueMatches2 = regex.search(ur'(\d\d\d\d)s', unicode(value))
                    valueMatches3 = regex.search(ur'(\d\d\d\d).{0,1}(\d\d)', unicode(value))
                    if valueMatches != None:
                        valueGroups = valueMatches.groups()

                        if valueGroups[0] != None:
                            obj.append( { 'value' : int(valueGroups[0]), 'avalue': annotationValue, 'type' : 'startYear', 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )
                        if valueGroups[1] != None:
                            obj.append( { 'value' : int(valueGroups[1]), 'avalue': annotationValue, 'type' : 'endYear', 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )
                    elif valueMatches3 != None:
                        valueGroups = valueMatches3.groups()

                        if valueGroups[0] != None:
                            obj.append( { 'value' : int(valueGroups[0]), 'avalue': annotationValue, 'type' : 'startYear', 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )
                        if valueGroups[1] != None:
                            obj.append( { 'value' : int(valueGroups[0][:2] + valueGroups[1]), 'avalue': annotationValue, 'type' : 'endYear', 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )


                    elif valueMatches2 != None:
                        obj.append( { 'value' : int(valueMatches2.groups()[0]), 'avalue': annotationValue, 'type' : 'startYear', 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )
                        obj.append( { 'value' : int(valueMatches2.groups()[0]), 'avalue': annotationValue, 'type' : 'endYear', 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )

                    else:
                        #obj.append( { 'value' : value, 'avalue': annotationValue, 'type' : annotationType, 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )
                        obj.append( { 'value' : value, 'avalue': annotationValue, 'type' : 'startYear', 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )
                        obj.append( { 'value' : value, 'avalue': annotationValue, 'type' : 'endYear', 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )
                    # link - mapping to URIRef
                else:
                    obj.append( { 'value' : value, 'avalue': annotationValue, 'type' : annotationType, 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex} )

                self.cells[colIndex][rowIndex] = obj
                if(isValue):
                    # check for multivalue
                    valueMatches = combinedNumberRegex.match(unicode(value).strip())
                    if valueMatches != None:
                        valueGroups = valueMatches.groups()
                        if len(valueGroups) > 1:
                            # create additional value object for the second match
                            #if valueGroups[1] != None: # was 2
                            #    self.values.append({ 'value' : float(valueGroups[1]), 'avalue': int(annotationValue), 'type' : annotationType, 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex})
                            #else:
                            self.values.append({ 'value' : int(valueGroups[0]), 'avalue': 1, 'type' : annotationType, 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex})
                            self.values.append({ 'value' : float(valueGroups[1]), 'avalue': int(annotationValue), 'type' : annotationType, 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex})
                    else:
                        self.values.append({ 'value' : value, 'avalue': annotationValue, 'type' : annotationType, 'isValue': isValue, 'isURI': isURI, 'isCommon': isCommon, 'rowIndex': rowIndex, 'colIndex': colIndex})
                if(isCommon):
                    self.common.extend(obj)

                if cell.style.startswith('0:Dataset:'):
                    self.dataset.extend(obj)

        #print self.cells[2][7]
        #print "value = " + unicode(self.cells[2][7]['value']).encode('utf-8', 'ignore')



    def initSheet(self, sheet):
        self.traverseCells(sheet)
        self.sheet = sheet

    def addDimension(self, g, subject, obj):
        #print obj
        if not(obj['avalue'].startswith('0:Dataset')):
            # only one dimension each allowed - for now
            # if it not common
            #if obj['type'] == 'Corpus' :
            if (subject, URIRef(baseURL + obj['type']), None) in  g and not(obj['isCommon']):
                return

            if obj['isURI']:
                uriValue = obj['value']
                if not uriValue.startswith('http://'):
                    uriValue = baseURL + obj['value']
                g.add( (subject, URIRef(baseURL + obj['type']), URIRef(uriValue) ))
            else:
                v = obj['value']
                if type(v) == str:
                    v = v.strip()
                g.add((subject, URIRef(baseURL + obj['type']), Literal(v) ))


    def isDimension(self, obj):
        return obj['type'] != 'Normal' and obj['type'] != 'Value' and not(obj['isCommon'])

    def generateGraph(self, prefix):
        g = Graph()
        # data set description
        datasetURI = URIRef( prefix + '-dataset-' + self.sheet.title )
        g.add( (datasetURI, RDF.type, qb_DataSet) )

        for value in self.dataset:
            style = value['avalue']
            parts = style.split(':')
            if parts[2] == 'file':
                g.add( (datasetURI, vg_file, URIRef(value['value'])))
            elif parts[2] == 'label':
                g.add( (datasetURI, RDFS.label, Literal(value['value'])))
            elif parts[2] == 'comment':
                g.add( (datasetURI, RDFS.comment, Literal(value['value'])))

        # Handle values
        for value in self.values:
            #print value
            #print "Value from " + str(value['colIndex']) + ", " + str(value['rowIndex'])
            valueCol = value['colIndex']
            valueRow = value['rowIndex']
            valueBase = value['avalue']

            subject = URIRef(':'  + prefix + str(valueCol) + '_' + str(valueRow) + '_' + str(valueBase))

            g.add( (subject, RDF.type, qb_obs ))
            g.add( (subject, qb_dataSet, datasetURI))
            valueStr = value['value']
            if valueStr == None:
                continue
            try:
                g.add( (subject, lada_row , Literal(int(valueRow))) )
                g.add( (subject, lada_col , Literal(int(valueCol))) )
                g.add( (subject, lada_sheet , Literal(self.sheet.title)) )


                v = float(valueStr)
                if v.is_integer():
                    g.add( (subject, vg_freq, Literal(int(v))) )
                else:
                    g.add( (subject, vg_freq, Literal(v)) )
            except ValueError:
                valueStr = valueStr.strip()
                if valueStr == '-':
                    g.add( (subject, vg_freq, Literal(0)) )
                else:
                    g.add( (subject, vg_freq, Literal(valueStr)) )

            try:
                g.add( (subject, vg_per, Literal(int(value['avalue'] ))))
            except:
                print "Error handled while add per value. Is the name of the style an integer?"
                g.add( (subject, vg_per, Literal(1)))
            # fixed row
            for x in range(1, self.maxCol):
                objs = self.cells[x][valueRow]
                for obj in objs:
                    if self.isDimension(obj):
                        self.addDimension(g, subject, obj)
                    for common in self.common:
                        self.addDimension(g, subject, common)
            # fixed col
            #for y in range(1, self.maxRow):
            for y in range(valueRow, 1, -1):
                objs = self.cells[valueCol][y]
                for obj in objs:
                    if self.isDimension(obj):
                        self.addDimension(g, subject, obj)
                    for common in self.common:
                        self.addDimension(g, subject, common)


        return g

    def isSkipped(self):
        return False
